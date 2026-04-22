"""마스터 스프레드시트 생성 (.xlsx)."""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.constants import (
    GRADE_FAIL,
    GRADE_IN_PROGRESS,
    GRADE_PASS,
    SHEET_ASSIGNMENT_DEF,
    SHEET_ASSIGNMENT_STATUS,
    SHEET_MASTER,
)

_HEADER_FILL = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_PASS_FILL = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="EA4335", end_color="EA4335", fill_type="solid")
_IN_PROGRESS_FILL = PatternFill(start_color="FBBC04", end_color="FBBC04", fill_type="solid")

_GRADE_FILLS = {
    GRADE_PASS: _PASS_FILL,
    GRADE_FAIL: _FAIL_FILL,
    GRADE_IN_PROGRESS: _IN_PROGRESS_FILL,
}


class ExcelExporter:
    def export(
        self,
        master_df: pd.DataFrame,
        assignment_defs: pd.DataFrame,
        assignment_status: pd.DataFrame,
        output_path: Path,
    ) -> Path:
        """3개 시트 + 요약 시트를 가진 .xlsx 파일 생성."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)  # 기본 시트 제거

        self._write_sheet(wb, SHEET_MASTER, master_df, grade_col="grade")
        self._write_sheet(wb, SHEET_ASSIGNMENT_DEF, assignment_defs)
        self._write_sheet(wb, SHEET_ASSIGNMENT_STATUS, assignment_status)
        self._add_summary_sheet(wb, master_df)

        wb.save(output_path)
        return output_path

    def _write_sheet(
        self,
        wb: Workbook,
        title: str,
        df: pd.DataFrame,
        grade_col: str | None = None,
    ) -> None:
        ws = wb.create_sheet(title=title)
        if df.empty:
            ws.append(["데이터 없음"])
            return

        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        grade_col_idx: int | None = None
        if grade_col and grade_col in df.columns:
            grade_col_idx = df.columns.tolist().index(grade_col) + 1

        for row_data in df.itertuples(index=False, name=None):
            ws.append(list(row_data))
            if grade_col_idx is not None:
                grade_val = row_data[grade_col_idx - 1]
                fill = _GRADE_FILLS.get(grade_val)
                if fill:
                    ws.cell(row=ws.max_row, column=grade_col_idx).fill = fill

        self._style_sheet(ws, df)

    def _style_sheet(self, ws, df: pd.DataFrame) -> None:
        for col_idx, col_name in enumerate(df.columns, start=1):
            lengths = [len(str(col_name))] + [
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(2, ws.max_row + 1)
            ]
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(lengths) + 4, 40)

    def _add_summary_sheet(self, wb: Workbook, master_df: pd.DataFrame) -> None:
        ws = wb.create_sheet(title="요약")

        def _header(text: str) -> None:
            ws.append([text])
            cell = ws.cell(row=ws.max_row, column=1)
            cell.font = Font(bold=True)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT

        _header("카테고리별 수강생 수")
        ws.append(["카테고리", "수강생 수"])
        if not master_df.empty and "category" in master_df.columns:
            for cat, cnt in master_df["category"].value_counts().items():
                ws.append([cat, cnt])
        ws.append([])

        _header("클래스별 이수율")
        ws.append(["클래스", "이수", "미이수", "진행중", "이수율(%)"])
        if not master_df.empty and "class_level" in master_df.columns and "grade" in master_df.columns:
            for cls, grp in master_df.groupby("class_level"):
                passed = (grp["grade"] == GRADE_PASS).sum()
                failed = (grp["grade"] == GRADE_FAIL).sum()
                in_prog = (grp["grade"] == GRADE_IN_PROGRESS).sum()
                total = len(grp)
                rate = round(passed / total * 100, 1) if total else 0.0
                ws.append([cls, passed, failed, in_prog, rate])
        ws.append([])

        _header("기수별 현황")
        ws.append(["기수", "수강생 수"])
        if not master_df.empty and "cohort" in master_df.columns:
            for cohort, cnt in master_df["cohort"].value_counts().sort_index().items():
                ws.append([cohort, cnt])

        for col_idx in range(1, 6):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
