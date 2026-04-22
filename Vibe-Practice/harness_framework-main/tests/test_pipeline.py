"""전체 파이프라인 통합 테스트 (Google Drive 없이 로컬 파일만 사용)."""

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from exporters.excel_exporter import ExcelExporter
from parsers.filename_parser import ParsedFileName
from processors.assignment_manager import AssignmentManager
from processors.grade_calculator import calculate_all
from processors.student_processor import StudentProcessor
from utils.constants import SHEET_MASTER


def _make_enrollment_xlsx(tmp_path: Path, filename: str = "test.xlsx") -> Path:
    """수강신청 응답 형식의 테스트 xlsx 파일 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "타임스탬프", "이메일 주소", "성함", "연락처", "섬기시는 교회",
        "직분", "수강 희망 클래스", "수강료 입금 여부", "개인정보 동의",
    ])
    ws.append([
        "2026/04/01 10:00:00", "alice@example.com", "홍길동", "010-1234-5678",
        "OO교회", "집사", "Class #1", "예", "예",
    ])
    ws.append([
        "2026/04/02 11:00:00", "bob@example.com", "김철수", "010-9876-5432",
        "AA교회", "권사", "Class #1", "아니오", "예",
    ])
    path = tmp_path / filename
    wb.save(path)
    return path


def _parsed_info(month: int = 4) -> ParsedFileName:
    return ParsedFileName(
        month=month,
        category="Wave Academy",
        target="일반",
        class_levels=[1],
        region=None,
        raw=f"(2026-{month:02d} 일반 Class#1)(응답)",
    )


def test_full_pipeline_local(tmp_path: Path) -> None:
    """로컬 xlsx → 처리 → 학점계산 → 내보내기 전체 흐름 검증."""
    # 1. 테스트용 수강신청 xlsx 생성
    enrollment_file = _make_enrollment_xlsx(tmp_path)

    # 2. StudentProcessor로 처리
    processor = StudentProcessor()
    rows = processor.process_file(enrollment_file, _parsed_info())
    assert rows == 2, f"처리된 행이 2여야 함 (실제: {rows})"

    master_df = processor.get_master()
    assert len(master_df) == 2

    # 3. GradeCalculator로 학점 계산
    graded_df = calculate_all(master_df, threshold=80.0)
    assert "grade" in graded_df.columns
    assert len(graded_df) == 2

    # 4. ExcelExporter로 내보내기
    am = AssignmentManager()
    output_path = tmp_path / "master_output.xlsx"
    exporter = ExcelExporter()
    result = exporter.export(
        master_df=graded_df,
        assignment_defs=am.get_definitions(),
        assignment_status=am.get_status(),
        output_path=output_path,
    )

    # 5. 결과 Excel 검증
    assert result.exists(), "결과 파일이 생성되어야 함"

    wb = openpyxl.load_workbook(result)
    assert SHEET_MASTER in wb.sheetnames, f"{SHEET_MASTER} 시트가 있어야 함"

    ws = wb[SHEET_MASTER]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    assert len(data_rows) == 2, f"데이터 행이 2여야 함 (실제: {len(data_rows)})"

    # 컬럼 확인
    headers = [c.value for c in ws[1]]
    for col in ("email", "name", "grade", "payment_status"):
        assert col in headers, f"컬럼 '{col}'이 있어야 함"


def test_pipeline_empty_file(tmp_path: Path) -> None:
    """빈 파일(헤더만) 처리 시 0행 반환."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["타임스탬프", "이메일 주소", "성함"])
    path = tmp_path / "empty.xlsx"
    wb.save(path)

    processor = StudentProcessor()
    rows = processor.process_file(path, _parsed_info())
    assert rows == 0


def test_pipeline_dedup(tmp_path: Path) -> None:
    """동일 email+class+cohort 중복 신청 → upsert (행 수 유지)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["타임스탬프", "이메일 주소", "성함", "연락처", "섬기시는 교회", "직분", "수강 희망 클래스", "수강료 입금 여부", "개인정보 동의"])
    ws.append(["2026/04/01 10:00:00", "dup@example.com", "중복", "010-0000-0000", "교회", "집사", "Class #1", "예", "예"])
    ws.append(["2026/04/02 10:00:00", "dup@example.com", "중복2", "010-0000-0000", "교회", "집사", "Class #1", "예", "예"])
    path = tmp_path / "dup.xlsx"
    wb.save(path)

    processor = StudentProcessor()
    processor.process_file(path, _parsed_info())
    master = processor.get_master()
    assert len(master) == 1, "중복 행은 upsert되어야 함"


def test_pipeline_stats(tmp_path: Path) -> None:
    """처리 통계(new/updated/errors) 정상 집계."""
    file1 = _make_enrollment_xlsx(tmp_path, "file1.xlsx")
    file2 = _make_enrollment_xlsx(tmp_path, "file2.xlsx")

    processor = StudentProcessor()
    processor.process_file(file1, _parsed_info())
    processor.process_file(file2, _parsed_info())

    stats = processor.get_stats()
    # file1에서 new 2, file2에서 updated 2 (동일 email+class+cohort)
    assert stats["new"] == 2
    assert stats["updated"] == 2
    assert stats["errors"] == 0
